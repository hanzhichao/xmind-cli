from pathlib import Path
import openpyxl
from ..core.models import Workbook, Topic

class ExcelConverter:
    @staticmethod
    def _get_paths(topic: Topic, current_path: list, current_level: int, start_level: int) -> list:
        # Only add to path if we have reached start_level
        if current_level >= start_level:
            path = current_path + [topic.title]
        else:
            path = current_path
            
        if not topic.children:
            return [path] if path else []
            
        all_paths = []
        for child in topic.children:
            all_paths.extend(ExcelConverter._get_paths(child, path, current_level + 1, start_level))
        return all_paths

    @classmethod
    def from_xmind(cls, workbook: Workbook, output_path: str | Path, headers: str = None, start_level: int = 1):
        output_path = Path(output_path)
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if wb.sheetnames:
            del wb[wb.sheetnames[0]]
            
        header_list = [h.strip() for h in headers.split(",")] if headers else []
            
        for sheet_data in workbook.sheets:
            ws = wb.create_sheet(title=sheet_data.title[:31]) # Excel limits title length to 31
            
            paths = cls._get_paths(sheet_data.root_topic, [], 1, start_level)
            
            # Write headers
            max_len = max(len(p) for p in paths) if paths else 1
            for col in range(1, max_len + 1):
                header_name = header_list[col - 1] if col - 1 < len(header_list) else f"Level {col}"
                ws.cell(row=1, column=col, value=header_name)
                
            # Write data
            for row_idx, path in enumerate(paths, start=2):
                for col_idx, value in enumerate(path, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
        wb.save(output_path)
